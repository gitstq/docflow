"""Excel to Markdown converter."""

from pathlib import Path

from docflow.core.converters.base import BaseConverter
from docflow.core.models import DocumentMetadata

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelConverter(BaseConverter):
    """Convert Excel spreadsheets to Markdown."""

    def convert(self, source_path: Path, output_path: Path):
        """Convert an Excel spreadsheet to Markdown."""
        if not HAS_OPENPYXL:
            return self._create_error_result(
                source_path,
                output_path,
                ["openpyxl not installed. Run: pip install openpyxl"],
            )

        try:
            wb = openpyxl.load_workbook(source_path, data_only=True)
            markdown_parts = []

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                # Apply sheet pattern filter if specified
                if self.options.excel_sheet_pattern:
                    import re
                    if not re.search(self.options.excel_sheet_pattern, sheet_name):
                        continue

                markdown_parts.append(f"\n\n## Sheet: {sheet_name}\n\n")

                # Get dimensions
                max_row = sheet.max_row
                max_col = sheet.max_column

                if max_row == 0 or max_col == 0:
                    markdown_parts.append("*Empty sheet*\n")
                    continue

                # Read data
                rows = []
                for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
                    cells = [str(cell.value) if cell.value is not None else "" for cell in row]
                    rows.append(cells)

                # Convert to Markdown table
                if rows:
                    md_table = self._format_rows(rows)
                    markdown_parts.append(md_table)

            markdown_content = "".join(markdown_parts).strip()
            metadata = self._extract_metadata(source_path, wb)

            return self._create_success_result(
                source_path,
                output_path,
                markdown_content,
                metadata=metadata,
            )

        except Exception as e:
            return self._create_error_result(
                source_path, output_path, [f"Excel conversion error: {str(e)}"]
            )

    def _format_rows(self, rows: list[list[str]]) -> str:
        """Format rows as a Markdown table."""
        if not rows:
            return ""

        header = rows[0]
        separator = ["---"] * len(header)
        body = rows[1:]

        lines = [
            "| " + " | ".join(header) + " |",
            "| " + " | ".join(separator) + " |",
        ]
        for row in body:
            # Pad row if necessary
            while len(row) < len(header):
                row.append("")
            lines.append("| " + " | ".join(row[:len(header)]) + " |")

        return "\n".join(lines)

    def _extract_metadata(self, source_path: Path, wb) -> DocumentMetadata:
        """Extract metadata from Excel."""
        props = wb.properties

        return DocumentMetadata(
            title=props.title,
            author=props.creator,
            subject=props.subject,
            keywords=props.keywords.split(", ") if props.keywords else [],
            creation_date=props.created,
            modification_date=props.modified,
            page_count=len(wb.sheetnames),
            file_size=source_path.stat().st_size if source_path.exists() else None,
        )
